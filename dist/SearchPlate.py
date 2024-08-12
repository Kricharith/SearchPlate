from tkinter import *
import tkinter as tk
from tkinter import ttk
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib
from datetime import datetime
import time
import pandas as pd
#background ="#FFF"

framebg="#ECEFF1" #พื้นหลังในกรอบ
framefg="#212121" #ตัวอักษรในกรอบ


root =Tk()
width= root.winfo_screenwidth()
height= root.winfo_screenheight()
root.geometry("%dx%d" % (width, height))

p1 = PhotoImage(file = 'Images/car.png')
# Setting icon of master window
root.iconphoto(False, p1)

#root.attributes('-fullscreen', True)
root.title("ระบบค้นหาข้อมูลเพลต")
#root.geometry("1920x1080+210+100")
#root.config(bg=background)
#Exit window
def Exit():
    root.destroy()
def Clear():
        Group.set("เลือก")
        conment.set("")  
        id_Oparator.set("")
        Search.set("")
        for widget in frame0.winfo_children():
            widget.destroy()
        for widget in frame_.winfo_children():
            widget.destroy()
        for widget in frame1.winfo_children():
            widget.destroy()
        for widget in frame12.winfo_children():
            widget.destroy()
        #listbox.delete(0,END)
def clock():
   hh= time.strftime("%I")
   mm= time.strftime("%M")
   ss= time.strftime("%S")
   my_lab.config(text= hh + ":" + mm +":" + ss)
   my_lab.after(1000,clock)

def dateget():
    Date = StringVar()
    today =date.today()
    d1 = today.strftime("%d/%m/%Y")
    Date.set(d1)
    return Date

def clock2():
   hh= time.strftime("%I")
   mm= time.strftime("%M")
   ss= time.strftime("%S")
   myTime= hh + ":" + mm +":" + ss 
   print(myTime)
   return myTime
def updateTime():
   my_lab.config(text= "New Text")

#Search
location =[]
def search():
    for widget in frame12.winfo_children():
        widget.destroy()
    list_column = [""]
    side = []
    side_data=[]
    location_data =[]
    text = Search.get()
    choicesA=[]
    choicesB=[]
    choicesC=[]
    choicesD=[]
    mark=[]
    markA=[]
    markB=[]
    markC=[]
    markD=[]
    note=""
    global listboxA
    global listboxB
    global listboxC
    global listboxD
    text = text.replace(" ", "")
    
    #file=openpyxl.load_workbook("ItemPad.xlsx", data_only=True)
    file=openpyxl.load_workbook("Location PlateJig.xlsx", data_only=True)
    #sheet = file['UPDATE 20-4-2023']
    sheet = file.worksheets[0]  #ชีทแรกสุดจากสมุดงานทั้งหมด
    sheet=file.active
    count = 0
    for row in sheet.rows:
        if str(row[2].value).replace(" ", "") == text:
            #print(text)
            #print("okkkkkkkkkkkk")
            name=row[2]
            item_pad_position=str(name)[15:-1]
            item_pad=str(name)[19:-1]    #กรองเอาแค่ เลขคอลัมพ์ เช่น <Worksheet "Data Plat"> C10 ....10
            #print(item_pad)
    try:
        for x in range(300):
            list_column.append(sheet.cell(row=int(item_pad),column=x+1).value)
        #print(list_column)    
        for x in range(4,300):
            if list_column[x] != None and str(sheet.cell(row=1,column=x).value) == "หมายเหตุ":
                #print("noteeeeeeeeeeeeeee")
                #print(list_column[x])
                note = list_column[x]
            if list_column[x] != None and list_column[x] == "*":
                mark.append("+"+str(sheet.cell(row=1,column=x-3).value))
            elif list_column[x] != None and list_column[x] == "-":
                mark.append("-"+str(sheet.cell(row=1,column=x-3).value))
            if list_column[x] != None and list_column[x] != "*" and list_column[x] != "-":
                count += 1
                side.append(list_column[x])
                if count == 1:
                    side_data.append(sheet.cell(row=1,column=x).value)
                if count == 3:
                    count = 0
                #print(count)
        # print(side)   
        # print(side_data)
        # print(mark)
        ########################################
        for datamark in range(len(mark)):
            if mark[datamark][0]=="+" and mark[datamark][1]=="A":
                markA.append("mark")
            elif mark[datamark][0]=="-" and mark[datamark][1]=="A":
                markA.append("Nomark")
            elif mark[datamark][0]=="+" and mark[datamark][1]=="B":
                markB.append("mark")
            elif mark[datamark][0]=="-" and mark[datamark][1]=="B":
                markB.append("Nomark")
            elif mark[datamark][0]=="+" and mark[datamark][1]=="C":
                markC.append("mark")
            elif mark[datamark][0]=="-" and mark[datamark][1]=="C":
                markC.append("Nomark")
            elif mark[datamark][0]=="+" and mark[datamark][1]=="D":
                markD.append("mark")
            elif mark[datamark][0]=="-" and mark[datamark][1]=="D":
                markD.append("Nomark")
        # print("mark")
        # print(markA)
        # print(markB)
        # print(markC)
        # print(markD)
        ########################################
        lengh = int(len(side))/3
        count2 = 0
        for i in range(int(lengh)):
            for n in range(3):
                count2 += 1
            srt_side_data = side_data[i]+" "*25
            str_side = str(side[count2-3])+"      "+str(side[count2-2])+"        "+str(side[count2-1])
            new_str_mix = srt_side_data[:10] + str_side + srt_side_data[25:]
            location_data.append(new_str_mix)
        #print(location_data)  
        side_data = location_data[:]
        choicesA.append("A             ตู้   คอลัมพ์   ช่อง")
        choicesB.append("B             ตู้   คอลัมพ์   ช่อง")
        choicesC.append("C             ตู้   คอลัมพ์   ช่อง")
        choicesD.append("D             ตู้   คอลัมพ์   ช่อง")
        for x in range(len(side_data)):
            if side_data[x][0]=="A":
                choicesA.append(side_data[x])
            elif side_data[x][0]=="B":
                choicesB.append(side_data[x])
            elif side_data[x][0]=="C":
                choicesC.append(side_data[x])
            elif side_data[x][0]=="D":
                choicesD.append(side_data[x])
        for widget in frame1.winfo_children():
            widget.destroy()
        for widget in frame_.winfo_children():
            widget.destroy()
        for widget in frame0.winfo_children():
            widget.destroy()
        Label(frame1,text="   ด้าน : ",fg='#212121',font='Tahoma 18 bold').pack(side = LEFT,padx= 10,pady=10,anchor="n")
        listboxA = ChecklistBox(frame1, choicesA,checkbuttonA,markA, bd=1, relief="sunken", background="white") 
        listboxA.pack( side = LEFT,padx= 10,anchor="nw",pady=10) 
        listboxB = ChecklistBox(frame1, choicesB,checkbuttonB,markB, bd=1, relief="sunken", background="white") 
        listboxB.pack( side = LEFT,padx= 10,anchor="nw",pady=10)  
        listboxC = ChecklistBox(frame1, choicesC,checkbuttonC,markC, bd=1, relief="sunken", background="white") 
        listboxC.pack( side = LEFT,padx= 10,anchor="nw",pady=10)
        listboxD = ChecklistBox(frame1, choicesD,checkbuttonD,markD, bd=1, relief="sunken", background="white") 
        listboxD.pack( side = LEFT,padx= 10,anchor="nw",pady=10)
        Label(frame12,text="  หมายเหตุ : ",fg='#212121',font='Tahoma 16').pack(side = LEFT,padx= 4,pady=4)
        Label(frame12,text=note,fg='#212121',font='Tahoma 16').pack(side = LEFT,padx= 4,pady=4)
    except: 
        messagebox.showerror("Invalid","Invalid Item Pad!!!")   

def savePlate():
    # print("choicesA:", listboxA.getCheckedItems())
    # print("choicesB:", listboxB.getCheckedItems())
    # print("choicesC:", listboxC.getCheckedItems())
    # print("choicesD:", listboxD.getCheckedItems())
    # print(len(listboxA.getCheckedItems()))
    # print(len(listboxB.getCheckedItems()))
    # print(len(listboxC.getCheckedItems()))
    # print(len(listboxD.getCheckedItems()))
    list_plate=[]
    savePlate=''
    for i in range(len(listboxA.getCheckedItems())):
        if listboxA.getCheckedItems()[i] == "A             ตู้   คอลัมพ์   ช่อง":
            pass
        else:
            list_plate.append(listboxA.getCheckedItems()[i].replace(" ", ""))
    for i in range(len(listboxB.getCheckedItems())):
        if listboxB.getCheckedItems()[i] == "B             ตู้   คอลัมพ์   ช่อง":
            pass
        else:
            list_plate.append(listboxB.getCheckedItems()[i].replace(" ", ""))
    for i in range(len(listboxC.getCheckedItems())):
        if listboxC.getCheckedItems()[i] == "C             ตู้   คอลัมพ์   ช่อง":
            pass
        else:
            list_plate.append(listboxC.getCheckedItems()[i].replace(" ", ""))
    for i in range(len(listboxD.getCheckedItems())):
        if listboxD.getCheckedItems()[i] == "D             ตู้   คอลัมพ์   ช่อง":
            pass
        else:
            list_plate.append(listboxD.getCheckedItems()[i].replace(" ", ""))
    for n in range(len(list_plate)):
        plate = str(list_plate[n])
        savePlate = savePlate+plate+","
        # print(plate)
        # print(savePlate)
    return savePlate


def check_Id():
    #print("check Id")
    df = pd.read_excel('Attendance PAD 2023.xlsx', sheet_name=None,skiprows=2)
    #print(df)
    #print(df.keys())
    #df['Pad A'].head()
    #print(df['Pad A'].head())
    try:
        id_op = id_Oparator.get()
        #print(id_op)
        if id_op[0] == '0':
            id_op = id_op.replace('0','', 1)
        id_op = float(id_op)
        #print(id_op)
    except:
        messagebox.showerror("ผิดพลาด","เลขIDไม่ถูกต้อง!!!")
    try:
        df2=df["Pad A"].query("ID == @id_op")
        #print(df2.index[0])
        if df2.index[0] == NONE :
            print("Error")    
        #print(df2)
        #print(df2['Name'])
        firstName = (df2['Name'])
        #print("(df2['Name']).index:")
        #print((df2['Name']).index[0])
        firstName = firstName[firstName.index[0]]
        lastName = (df2['Unnamed: 3'])
        lastName = lastName[lastName.index[0]]
        #print(firstName)
        #print(lastName)  
        openNewWindow(firstName,lastName)
    except:
        messagebox.showerror("ผิดพลาด","เลขIDไม่ถูกต้อง!!!")
    
window_height = 360
window_width = 500
def openNewWindow(firstName,lastName):
    newWindow = Toplevel(root)
    newWindow.title("ยืนยันการเบิกของ")
    newWindow.resizable(width=0,height=0)
    center_screen(newWindow)

    id_op = id_Oparator.get()
    Label(newWindow,text="รหัสพนักงาน :",fg='#212121',font='Tahoma 12 bold').place(x=5,y=10)
    Label(newWindow,text=id_op,fg='#212121',font='Tahoma 12 bold').place(x=125,y=10)

    Label(newWindow,text="ชื่อ :",fg='#212121',font='Tahoma 12 bold').place(x=5,y=50)
    Label(newWindow,text=firstName,fg='#212121',font='Tahoma 12 bold').place(x=45,y=50)

    Label(newWindow,text="นามสกุล :",fg='#212121',font='Tahoma 12 bold').place(x=140,y=50)
    Label(newWindow,text=lastName,fg='#212121',font='Tahoma 12 bold').place(x=230,y=50)

    id_Pad = Search.get()
    Label(newWindow,text="ID Pad :",fg='#212121',font='Tahoma 12 bold').place(x=5,y=90)
    Label(newWindow,text=id_Pad,fg='#212121',font='Tahoma 12 bold').place(x=80,y=90)

    plate = savePlate()
    print(plate)
    print(len(plate))
    if len(plate)>=44:
        plate_group1 = plate[:44] + plate[len(plate):]
        plate_group2 = plate[:0] + plate[44:]
        print("plate_group1 :")
        print(plate_group1)
        print("plate_group2 :")
        print(plate_group2)
        Label(newWindow,text="ด้านที่เบิก:",fg='#212121',font='Tahoma 12 bold').place(x=5,y=130)
        Label(newWindow,text=plate_group1,fg='#212121',font='Tahoma 12 bold').place(x=85,y=130)
        Label(newWindow,text=plate_group2,fg='#212121',font='Tahoma 12 bold').place(x=85,y=170)
    else:
        Label(newWindow,text="ด้านที่เบิก:",fg='#212121',font='Tahoma 12 bold').place(x=5,y=130)
        Label(newWindow,text=plate,fg='#212121',font='Tahoma 12 bold').place(x=85,y=130)
    Dateget = dateget()
    Label(newWindow,text="วันที่:",fg='#212121',font='Tahoma 12 bold').place(x=5,y=210)
    Label(newWindow,textvariable=Dateget,fg='#212121',font='Tahoma 12 bold').place(x=50,y=210)

    time = clock2()
    Label(newWindow,text="เวลา:",fg='#212121',font='Tahoma 12 bold').place(x=160,y=210)
    Label(newWindow,text=time,fg='#212121',font='Tahoma 12 bold').place(x=210,y=210)

    location = Group.get()
    Label(newWindow,text="สถานที่นำไปใช้:",fg='#212121',font='Tahoma 12 bold').place(x=5,y=250)
    Label(newWindow,text=location,fg='#212121',font='Tahoma 12 bold').place(x=150,y=250)

    conmentSave = conment.get()
    Label(newWindow,text="ความคิดเห็น :",fg='#212121',font='Tahoma 12 bold').place(x=5,y=280)
    Label(newWindow,text=conmentSave,fg='#212121',font='Tahoma 12 bold').place(x=150,y=280)

    cancle=Button(newWindow,text="ยกเลิก",width=7,height=1,bg='#68ddfa',font="Tahoma 12",command=newWindow.destroy)   
    cancle.place(x=150,y=315)

    confirm=Button(newWindow,text="ยืนยัน",width=7,height=1,bg='#68ddfa',font="Tahoma 12",command = lambda: confirmJob(newWindow,Dateget,time,id_Pad,id_op,firstName,lastName,plate,location,conmentSave))   
    confirm.place(x=280,y=315)

def confirmJob(newWindow,Dateget,time,id_Pad,id_op,firstName,lastName,plate,location,conmentSave):
    #print("confirmJob")
    time_ =''
    time_ =str(time)
    Dateget_ =''
    Dateget_ =str(Dateget)
    # print(Dateget_)
    # print(time_)
    # print(id_Pad)
    # print(id_op)
    # print(firstName)
    # print(lastName)
    # print(plate)
    file=pathlib.Path('saveData.xlsx')
    if file.exists():
        pass
    else:
        file=Workbook()
        sheet=file.active
        sheet['A1']="Date"
        sheet['B1']="Time"
        sheet['C1']="id_Pad"
        sheet['D1']="id_op"
        sheet['E1']="FirstName"
        sheet['F1']="LastName"
        sheet['G1']="Plate"
        sheet['H1']="สถานที่นำไปใช้"
        sheet['I1']="ความคิดเห็น"
        file.save('saveData.xlsx')
    #save sheet
    if Dateget=="" or time=="" or id_Pad=="" or id_op=="" or firstName=="" or lastName=="" or plate=="" or location =="":
        messagebox.showerror("error","Few Data is missing!")
    else:
        file=openpyxl.load_workbook('saveData.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=date.today().strftime("%d/%m/%Y"))  
        sheet.cell(column=2,row=sheet.max_row,value=time_) 
        sheet.cell(column=3,row=sheet.max_row,value=id_Pad)
        sheet.cell(column=4,row=sheet.max_row,value=id_op)
        sheet.cell(column=5,row=sheet.max_row,value=firstName)
        sheet.cell(column=6,row=sheet.max_row,value=lastName)
        sheet.cell(column=7,row=sheet.max_row,value=plate)
        sheet.cell(column=8,row=sheet.max_row,value=location)
        sheet.cell(column=9,row=sheet.max_row,value=conmentSave)
        file.save(r'saveData.xlsx')
    newWindow.destroy()

def center_screen(newWindow):
	global screen_height, screen_width, x_cordinate, y_cordinate
	screen_width = newWindow.winfo_screenwidth()
	screen_height = newWindow.winfo_screenheight()
	x_cordinate = int((screen_width/2) - (window_width/2))
	y_cordinate = int((screen_height/2) - (window_height/2))
	newWindow.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate)) 

checkbuttonA = StringVar()
checkbuttonB = StringVar()
checkbuttonC = StringVar()
checkbuttonD = StringVar()
checkbuttonAll =StringVar()
class ChecklistBox(tk.Frame):
    def __init__(self, parent, choices,checkbuttonAll,markCheck, **kwargs):
        tk.Frame.__init__(self, parent, **kwargs)
        self.vars = []
        bg = self.cget("background")
        countmark = 0

        # print("=================")
        # print(markCheck)
        for choice in choices:
            if choice =="A             ตู้   คอลัมพ์   ช่อง" or choice =="B             ตู้   คอลัมพ์   ช่อง" or choice =="C             ตู้   คอลัมพ์   ช่อง" or choice =="D             ตู้   คอลัมพ์   ช่อง":
                var = tk.StringVar(value=choice)
                self.vars.append(var)
                cb0 = tk.Label(self,text=" "+choice+" ",font="Tahoma 16",fg='#212121')
                cb0.pack(side="top", fill="x", anchor="w")
                #cb0.deselect()
            else:
                # เพิ่มมาค
                if markCheck[countmark] == "mark":
                    bgmark = '#ffbeb1'
                else:
                    bgmark = bg
                var = tk.StringVar(value=choice)
                self.vars.append(var)
                cb = tk.Checkbutton(self, var=var, text=choice,background=bgmark,
                                onvalue=choice, offvalue="",
                                anchor="w", width=20,
                                relief="flat", highlightthickness=0,font="Tahoma 16"
                )
                countmark += 1
                cb.pack(side="top", fill="x", anchor="w")
                cb.deselect()
    def getCheckedItems(self):
        values = []
        for var in self.vars:
            value =  var.get()
            if value:
                values.append(value)
        return values
    
#Date
Label(root,text="Date:",font="Tahoma 16",fg='#212121').place(x=5,y=5)
Date = StringVar()

today =date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Label(root,textvariable=Date,font="Tahoma 16",fg='#212121')
date_entry.place(x=80,y=5)
Date.set(d1)

Label(root,text= "Time: ",font="Tahoma 16", fg='#212121').place(x=5,y=35)
my_lab= Label(root,text= "",font="Tahoma 16", fg='#212121')
my_lab.place(x=80,y=35)
# my_lab1= Label(root, text= "",font="Tahoma 16",fg='#212121')
# my_lab1.place(x=120,y=35)
clock()

Label(root,text="",width=5,height=2,fg='#212121',font='Tahoma 15 bold').pack()

img =(Image.open("Images/majorette.png"))
resized_image=img.resize((450,200))
photo = ImageTk.PhotoImage(resized_image)
lbl=Label(image=photo)
lbl.pack( side = TOP)

#Label(root,text="ระบบค้นหาเพลต",width=30,height=2,fg='#212121',font='Tahoma 25 bold').place(x=500,y=250)

frame = Frame(root)
frame.pack()

imageicon3=PhotoImage(file="Images/majorette.png")

Label(frame,text="ค้นหา ITEM PAD",height=2,fg='#212121',font='Tahoma 20').pack( side = LEFT ,padx= 5,pady=15)
Search = StringVar()
Entry(frame,textvariable=Search,width=20,bd=2,font="CmPrasanmit 25").pack( side = LEFT)

imageicon3=PhotoImage(file="Images/Search2.png")
Srch=Button(frame,text="Search",compound=LEFT,image=imageicon3,width=123,height=38,bg='#68ddfa',font="Tahoma 14",command=search)
Srch.pack( side = LEFT ,padx= 5,pady=15)

Reset=Button(frame,text="Reset",width=6,height=1,bg='#68ddfa',font="Tahoma 16",command=Clear)
Reset.pack( side = LEFT ,padx= 5,pady=15)

frame0 = Frame(root)
frame0.pack()

frame_ = Frame(root)
frame_.pack()

frame1 = Frame(root)
frame1.pack()

frame12 = Frame(root)
frame12.pack()

frame2 = Frame(root)
frame2.pack()

frame3 = Frame(root)
frame3.pack()

id_Oparator = StringVar()
conment = StringVar()
timeOut = StringVar()
Label(frame2,text="รหัสพนักงาน : ",width=10,height=2,fg='#212121',font='Tahoma 12 bold').pack( side = LEFT)
Entry(frame2,textvariable=id_Oparator,width=10,bd=2,font="Tahoma 15").pack( side = LEFT,padx= 20)
     
Label(frame2,text=" สถานที่นำไปใช้ : ",width=14,height=2,fg='#212121',font='Tahoma 12 bold').pack( side = LEFT)
Group=Combobox(frame2,values=['Pad','IE',],font="Tahoma 12 ",width=17,state="r")
Group.pack(side = LEFT)
Group.set("เลือก")
Label(frame3,text="แสดงความคิดเห็น : ",width=15,height=2,fg='#212121',font='Tahoma 12 bold').pack( side = LEFT)
Entry(frame3,textvariable=conment,width=40,bd=2,font="Tahoma 15").pack( side = LEFT)

Reset=Button(root,text="เบิกงาน",width=7,height=1,bg='#68ddfa',font="Tahoma 16",command=check_Id)
Reset.pack()

root.mainloop()
